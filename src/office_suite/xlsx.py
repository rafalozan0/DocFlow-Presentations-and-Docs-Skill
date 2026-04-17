import os
from typing import Any, Dict, List, Optional

import openpyxl
import pandas as pd
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def create_excel(
    title: str,
    data: List[List[Any]],
    output_path: str,
    create_chart: bool = False,
    **kwargs,
) -> Dict[str, Any]:
    """Create an Excel workbook.

    Args:
        title: Worksheet title.
        data: 2D list data (first row is header).
        output_path: Destination file path.
        create_chart: Whether to add a simple bar chart.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:31] if title else "Sheet1"

    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for row_idx, row_data in enumerate(data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
            else:
                cell.alignment = left_align if col_idx == 1 else center_align
            cell.border = thin_border

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = min(adjusted_width, 50)

    if create_chart and len(data) >= 2 and len(data[0]) >= 2:
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = title
        chart.y_axis.title = "Value"
        chart.x_axis.title = str(data[0][0])

        data_ref = Reference(ws, min_col=2, min_row=1, max_row=len(data), max_col=len(data[0]))
        categories = Reference(ws, min_col=1, min_row=2, max_row=len(data))
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(categories)

        ws.add_chart(chart, f"A{len(data) + 3}")

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)

    return {
        "output_path": output_path,
        "file_size": os.path.getsize(output_path),
        "rows": len(data),
        "columns": len(data[0]) if data else 0,
    }


def extract_data(input_path: str, sheet_name: Optional[str] = None, **kwargs) -> List[Dict[str, Any]]:
    """Extract tabular data from an Excel sheet.

    Returns a list of dictionaries where keys are column names.
    """
    target_sheet = sheet_name if sheet_name is not None else 0
    df = pd.read_excel(input_path, sheet_name=target_sheet)
    return df.to_dict("records")


def from_dataframe(df: pd.DataFrame, output_path: str, **kwargs) -> Dict[str, Any]:
    """Create and format an Excel file from a pandas DataFrame."""
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    df.to_excel(output_path, index=False, **kwargs)

    wb = openpyxl.load_workbook(output_path)
    ws = wb.active

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[column].width = max_length + 2

    wb.save(output_path)

    return {"output_path": output_path, "rows": len(df), "columns": len(df.columns)}
